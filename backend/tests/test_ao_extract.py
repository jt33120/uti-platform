"""
Extraction XLSX pour la génération d'AO.

Sur un modèle de marché (type AGIRC-ARRCO : feuilles CCTP/CRT/AF + feuilles de
listes déroulantes), les feuilles « catalogue » volumineuses doivent être
reléguées en annexe pour que les vrais champs de la consultation arrivent en
tête du texte transmis à l'IA.
"""
import io

import openpyxl

from services.cv_parser import extract_text_from_xlsx


def _make_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    # 1re feuille = gros catalogue de référence (doit finir en annexe)
    cat = wb.active
    cat.title = "Liste Categorie Uos"
    for i in range(60):
        cat.append([f"UO{i}", f"Cat - UO{i} libellé générique de catalogue"])
    # 2e feuille = vraies données de la consultation
    cctp = wb.create_sheet("CCTP")
    cctp.append(["Références de la consultation :", "Marché Spécifique n°23915SA240MS"])
    cctp.append(["Objet de la consultation :", "Prestation d'administration plan de charge SENIOR"])
    cctp.append(["Date de limite de remise des offres:", "14/08/2026"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_reference_sheets_relegated_to_annex():
    txt = extract_text_from_xlsx(_make_xlsx())
    ref_marker = txt.find("[Listes de référence")
    consult = txt.find("Marché Spécifique n°23915SA240MS")
    assert consult >= 0, "la donnée de consultation doit être extraite"
    assert ref_marker >= 0, "le bloc d'annexes doit être présent"
    # Les vraies données précèdent le catalogue de référence.
    assert consult < ref_marker
    # Le titre de la feuille catalogue n'apparaît qu'après le marqueur d'annexe.
    assert txt.find("Liste Categorie Uos") > ref_marker


def test_reference_sheet_is_truncated():
    # Catalogue volumineux → tronqué à _XLSX_REF_MAX_CHARS dans l'annexe.
    wb = openpyxl.Workbook()
    cat = wb.active
    cat.title = "Liste déroulante"
    for i in range(2000):
        cat.append([f"valeur de liste déroulante numéro {i} avec du texte de remplissage"])
    buf = io.BytesIO()
    wb.save(buf)
    txt = extract_text_from_xlsx(buf.getvalue())
    assert len(txt) < 4000, "une feuille de listes déroulantes doit être tronquée"
